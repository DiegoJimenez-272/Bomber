from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.db.models import Sum, Case, When, DecimalField, F, Q, Count
from django.utils import timezone
from django.db.models.functions import TruncDay
from datetime import datetime, timedelta
import json, re
from io import BytesIO
from django.core.files.base import ContentFile
from django.http import HttpResponse, JsonResponse
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
import openpyxl
from reportlab.lib.units import inch
from .forms import RegistroForm, LoginForm, ProyectoForm, ArchivoProyectoForm, DocumentoForm, SalidaTerrenoForm, EmergenciaForm, PerfilForm, PasswordChangeForm, CapacitacionForm, MantenimientoForm, ArchivoMantenimientoForm, InventarioForm, CajaChicaForm, AdminUserCreationForm, AdminUserChangeForm, RolForm, CompaniaForm, InventarioEditForm, InventarioGroupEditForm
from .models import Rol, Compania, Proyecto, ArchivoProyecto, Documento, SalidaTerreno, Emergencia, Usuario, Capacitacion, Mantenimiento, ArchivoMantenimiento, Inventario, CajaChica, Notificacion

def registro_view(request):
    if request.method == 'POST':
        form = RegistroForm(request.POST, request.FILES)
        if form.is_valid():
            # Guardamos el usuario sin confirmarlo en la BD para poder modificarlo
            user = form.save(commit=False)
            # Marcamos la cuenta como inactiva hasta que un admin la apruebe
            user.is_active = False
            
            # Asignar rol por defecto (puedes crear uno en admin)
            rol_usuario, created = Rol.objects.get_or_create(
                nombre='Usuario',
                defaults={'descripcion': 'Usuario estándar del sistema'}
            )
            user.rol = rol_usuario
            user.save() # Ahora guardamos el usuario con los cambios
            
            # Enviar notificación al sistema para todos los administradores (superusers)
            superusers = Usuario.objects.filter(is_superuser=True)
            notificaciones = []
            for admin in superusers:
                notificaciones.append(Notificacion(
                    usuario=admin,
                    mensaje=f"Nueva solicitud de registro: {user.get_full_name()}",
                    link="/administracion/?tab=requests"
                ))
            Notificacion.objects.bulk_create(notificaciones)
            
            messages.success(request, '¡Registro exitoso! Tu cuenta ha sido creada y está pendiente de aprobación por un administrador.')
            return redirect('login')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = RegistroForm()
    
    return render(request, 'usuarios/registro.html', {'form': form})

def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            email = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, email=email, password=password)
            
            if user is not None:
                login(request, user)
                messages.success(request, f'¡Bienvenido {user.get_full_name()}!')
                return redirect('dashboard')
            else:
                messages.error(request, 'Credenciales inválidas')
        else:
            messages.error(request, 'Por favor corrige los errores')
    else:
        form = LoginForm()
    
    return render(request, 'usuarios/login.html', {'form': form})

def logout_view(request):
    logout(request)
    messages.info(request, 'Has cerrado sesión exitosamente.')
    return redirect('login')

@login_required
def dashboard_view(request):
    modules = [
        {'id': 'dashboard_summary', 'url_name': 'dashboard_summary', 'icon': 'bi-bar-chart-line-fill', 'title': 'Ver Resumen General', 'desc': 'Visualiza KPIs, gráficos y actividad reciente del sistema.'},
        {'id': 'documentos', 'url_name': 'documentos', 'icon': 'bi-file-earmark-text-fill', 'title': 'Documentos', 'desc': 'Gestión de documentos e informes'},
        {'id': 'inventario', 'url_name': 'inventario', 'icon': 'bi-box-seam-fill', 'title': 'Inventario', 'desc': 'Control de equipos y herramientas'},
        {'id': 'salidas', 'url_name': 'salidas_terreno', 'icon': 'bi-truck-front-fill', 'title': 'Salidas a Terreno', 'desc': 'Registro y seguimiento de operaciones'},
        {'id': 'emergencias', 'url_name': 'emergencias', 'icon': 'bi-exclamation-octagon-fill', 'title': 'Emergencias', 'desc': 'Gestión de incidentes y respuesta'},
        {'id': 'capacitaciones', 'url_name': 'capacitaciones', 'icon': 'bi-mortarboard-fill', 'title': 'Cursos', 'desc': 'Programación y seguimiento de cursos y entrenamientos'},
        {'id': 'mantenimiento', 'url_name': 'mantenimiento', 'icon': 'bi-wrench-adjustable-circle-fill', 'title': 'Mantenimiento', 'desc': 'Gestión de la flota vehicular'},
        {'id': 'caja_chica', 'url_name': 'caja_chica', 'icon': 'bi-wallet-fill', 'title': 'Caja Chica', 'desc': 'Control de gastos menores'},
        {'id': 'proyectos', 'url_name': 'proyectos', 'icon': 'bi-graph-up-arrow', 'title': 'Proyectos', 'desc': 'Seguimiento de proyectos y mejoras'},
        {'id': 'scan_qr', 'url_name': 'scan_qr_page', 'icon': 'bi-qr-code-scan', 'title': 'Escanear QR', 'desc': 'Identifica un ítem de inventario usando su código QR.'},
        {'id': 'administracion', 'url_name': 'administracion', 'icon': 'bi-gear-fill', 'title': 'Administración', 'desc': 'Configuración y gestión de usuarios'},
    ]

    # Filtrar módulos basados en los permisos del usuario
    if request.user.is_superuser:
        filtered_modules = modules
    else:
        filtered_modules = []
        rol = request.user.rol
        for m in modules:
            # Modulos estrictamente para superusuario
            if m['id'] in ['administracion']:
                continue
            
            # Módulos controlados mediante los permisos configurados en el Rol
            if rol:
                if m['id'] == 'dashboard_summary' and rol.ver_resumen_general: filtered_modules.append(m)
                elif m['id'] == 'documentos' and rol.ver_documentacion: filtered_modules.append(m)
                elif m['id'] == 'inventario' and rol.ver_inventario: filtered_modules.append(m)
                elif m['id'] == 'salidas' and rol.ver_salidas_terreno: filtered_modules.append(m)
                elif m['id'] == 'emergencias' and rol.ver_emergencias: filtered_modules.append(m)
                elif m['id'] == 'capacitaciones' and rol.ver_capacitaciones: filtered_modules.append(m)
                elif m['id'] == 'mantenimiento' and rol.ver_mantenimientos: filtered_modules.append(m)
                elif m['id'] == 'caja_chica' and rol.ver_caja_chica: filtered_modules.append(m)
                elif m['id'] == 'scan_qr' and rol.ver_escanear_qr: filtered_modules.append(m)
                elif m['id'] == 'proyectos' and rol.ver_proyectos: filtered_modules.append(m)

    context = {
        'modules': filtered_modules,
    }
    return render(request, 'usuarios/inicio.html', context)

@login_required
def dashboard_summary_view(request):
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.ver_resumen_general):
        messages.error(request, 'No tienes permiso para acceder al Resumen General.')
        return redirect('dashboard')

    # --- KPIs / Tarjetas de Resumen ---
    total_users = Usuario.objects.count()
    total_budget = Proyecto.objects.aggregate(total=Sum('presupuesto'))['total'] or 0
    emergencies_this_month = Emergencia.objects.filter(fecha_hora_alarma__month=timezone.now().month).count()

    # --- Lógica de Filtro para Caja Chica ---
    caja_chica_qs = CajaChica.objects.all()
    nombre_caja = "Global" # Título por defecto
    filtro_compania_caja = 'todas'

    if request.user.is_superuser:
        filtro_compania_caja = request.GET.get('compania_caja', 'todas')
        if filtro_compania_caja == 'general':
            caja_chica_qs = caja_chica_qs.filter(compania__isnull=True)
            nombre_caja = "Caja General"
        elif filtro_compania_caja != 'todas' and filtro_compania_caja.isdigit():
            caja_chica_qs = caja_chica_qs.filter(compania_id=filtro_compania_caja)
            try:
                nombre_caja = Compania.objects.get(id=filtro_compania_caja).nombre
            except Compania.DoesNotExist:
                nombre_caja = "Desconocida"
        else:
            nombre_caja = "Todas las Cajas"
    else: # Usuario no-admin
        filtro_compania_caja = None
        if request.user.compania:
            caja_chica_qs = caja_chica_qs.filter(compania=request.user.compania)
            nombre_caja = request.user.compania.nombre
        else:
            caja_chica_qs = CajaChica.objects.none()
            nombre_caja = "Sin Compañía"

    saldo_caja_chica = caja_chica_qs.aggregate(
        saldo=Sum(Case(When(tipo='Ingreso', then=F('monto')), When(tipo='Egreso', then=-F('monto')), output_field=DecimalField()))
    )['saldo'] or 0.00
    # --- Datos para Gráfico de Emergencias (Dona) ---
    emergencias_por_tipo = Emergencia.objects.values('tipo').annotate(count=Count('id')).order_by('-count')
    emergencia_labels = [item['tipo'] for item in emergencias_por_tipo]
    emergencia_data = [item['count'] for item in emergencias_por_tipo]

    # --- Datos para Gráfico de Caja Chica (Líneas) ---
    today = timezone.now()
    start_date = (today - timedelta(days=29)).replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Generar todas las fechas de los últimos 30 días
    date_range = [start_date + timedelta(days=x) for x in range(30)]
    caja_chica_labels = [d.strftime('%d/%m') for d in date_range]

    # Obtener datos de ingresos y egresos
    ingresos_data = caja_chica_qs.filter(tipo='Ingreso', fecha__gte=start_date) \
        .annotate(day=TruncDay('fecha')).values('day').annotate(total=Sum('monto')).order_by('day')
    egresos_data = caja_chica_qs.filter(tipo='Egreso', fecha__gte=start_date) \
        .annotate(day=TruncDay('fecha')).values('day').annotate(total=Sum('monto')).order_by('day')

    # Mapear datos a los días correspondientes
    ingresos_map = {item['day'].date(): item['total'] for item in ingresos_data}
    egresos_map = {item['day'].date(): item['total'] for item in egresos_data}

    caja_chica_ingresos = [float(ingresos_map.get(d.date(), 0)) for d in date_range]
    caja_chica_egresos = [float(egresos_map.get(d.date(), 0)) for d in date_range]

    # --- Actividad Reciente ---
    latest_emergencies = Emergencia.objects.all().order_by('-fecha_hora_alarma')[:5]
    latest_maintenances = Mantenimiento.objects.all().order_by('-fecha')[:5]
    latest_salidas = SalidaTerreno.objects.all().order_by('-fecha_hora_salida')[:5]
    
    # Combinar y ordenar actividades
    recent_activity = []
    for item in latest_emergencies:
        recent_activity.append({'type': 'Emergencia', 'obj': item, 'date': item.fecha_hora_alarma})
    for item in latest_maintenances:
        # Convertimos la fecha (date) a datetime para poder comparar
        if item.fecha:
            naive_datetime = datetime.combine(item.fecha, datetime.min.time()) # type: ignore
            date_as_datetime = timezone.make_aware(naive_datetime, timezone.get_default_timezone()) # type: ignore
            recent_activity.append({'type': 'Mantenimiento', 'obj': item, 'date': date_as_datetime})
    for item in latest_salidas:
        recent_activity.append({'type': 'Salida a Terreno', 'obj': item, 'date': item.fecha_hora_salida})

    # Ordenar por fecha (más reciente primero) y tomar los últimos 5
    recent_activity = sorted(recent_activity, key=lambda x: x['date'], reverse=True)[:5]

    # --- Datos para Gráfico de Proyectos ---
    proyectos_por_tipo = Proyecto.objects.values('tipo').annotate(count=Count('id')).order_by('-count')
    proyecto_labels = [item['tipo'] for item in proyectos_por_tipo]
    proyecto_data = [item['count'] for item in proyectos_por_tipo]

    # --- Datos para Gráfico de Inventario ---
    inventario_qs_agrupado = Inventario.objects.all()
    if not request.user.is_superuser and request.user.compania:
        inventario_qs_agrupado = inventario_qs_agrupado.filter(Q(compania=request.user.compania) | Q(compania__isnull=True))
    
    inventario_por_estado = inventario_qs_agrupado.values('estado').annotate(count=Count('id')).order_by('-count')
    inventario_labels = [item['estado'] for item in inventario_por_estado]
    inventario_data = [item['count'] for item in inventario_por_estado]

    # --- Próximos Cursos ---
    upcoming_courses = Capacitacion.objects.filter(fecha_inicio__gte=timezone.now()).order_by('fecha_inicio')[:5]

    # Compañías para el filtro (solo para superuser)
    companias_para_filtro = Compania.objects.all() if request.user.is_superuser else None

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({
            'saldo_caja_chica': float(saldo_caja_chica),
            'nombre_caja': nombre_caja,
            'caja_chica_labels': caja_chica_labels,
            'caja_chica_ingresos': caja_chica_ingresos,
            'caja_chica_egresos': caja_chica_egresos,
        })

    context = {
        'total_users': total_users,
        'total_budget': total_budget,
        'emergencies_this_month': emergencies_this_month,
        'saldo_caja_chica': saldo_caja_chica,
        'emergencia_labels': json.dumps(emergencia_labels),
        'emergencia_data': json.dumps(emergencia_data),
        'caja_chica_labels': json.dumps(caja_chica_labels),
        'caja_chica_ingresos': json.dumps(caja_chica_ingresos),
        'caja_chica_egresos': json.dumps(caja_chica_egresos),
        'proyecto_labels': json.dumps(proyecto_labels),
        'proyecto_data': json.dumps(proyecto_data),
        'inventario_labels': json.dumps(inventario_labels),
        'inventario_data': json.dumps(inventario_data),
        'recent_activity': recent_activity,
        'upcoming_courses': upcoming_courses,
        'companias_para_filtro': companias_para_filtro,
        'filtro_compania_caja': filtro_compania_caja,
        'nombre_caja': nombre_caja,
    }
    return render(request, 'usuarios/dashboard_summary.html', context)

@login_required
def generar_reporte_anual_pdf(request):
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permiso para generar este tipo de reportes.')
        return redirect('dashboard_summary')

    # --- Módulos a exportar ---
    modulos_str = request.GET.get('modulos', 'documentos,inventario,salidas,emergencias,cursos,mantenimientos,proyectos,caja_chica')
    modulos = [m.strip() for m in modulos_str.split(',') if m.strip()]

    # --- Configuración del Documento PDF ---
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_anual_{timezone.now().year}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter, topMargin=inch, bottomMargin=inch)
    story = []
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='TitleStyle', fontSize=24, alignment=TA_CENTER, spaceAfter=20, fontName='Helvetica-Bold'))
    styles.add(ParagraphStyle(name='HeaderStyle', fontSize=16, alignment=TA_LEFT, spaceAfter=12, fontName='Helvetica-Bold'))
    styles.add(ParagraphStyle(name='BodyStyle', fontSize=10, alignment=TA_LEFT, leading=14))

    # --- Contenido del PDF ---
    # Título
    story.append(Paragraph(f"Memoria Anual {timezone.now().year}", styles['TitleStyle']))
    story.append(Paragraph(f"Cuerpo de Bomberos - Reporte General del Sistema", styles['h2']))
    story.append(Spacer(1, 0.5 * inch))

    # Función para crear tablas estilizadas
    def create_table(data):
        table = Table(data, colWidths=[doc.width/len(data[0])]*len(data[0]))
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc3545')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        return table

    # 1. Proyectos
    if 'proyectos' in modulos:
        story.append(Paragraph("Resumen de Proyectos", styles['HeaderStyle']))
        proyectos = Proyecto.objects.all()
        proyectos_data = [['Nombre', 'Tipo', 'Responsable', 'Presupuesto']]
        for p in proyectos:
            proyectos_data.append([
                p.nombre, p.tipo, p.responsable.get_full_name() if p.responsable else 'N/A', f"${p.presupuesto or 0:,.0f}"
            ])
        story.append(create_table(proyectos_data))
        story.append(Spacer(1, 0.3 * inch))

    # 2. Emergencias
    if 'emergencias' in modulos:
        story.append(Paragraph("Resumen de Emergencias", styles['HeaderStyle']))
        emergencias = Emergencia.objects.all()
        emergencias_data = [['Tipo', 'Dirección', 'Fecha', 'Estado']]
        for e in emergencias:
            emergencias_data.append([
                e.get_tipo_display(), e.direccion, e.fecha_hora_alarma.strftime('%d/%m/%Y'), e.estado
            ])
        story.append(create_table(emergencias_data))
        story.append(Spacer(1, 0.3 * inch))

    # 3. Salidas a Terreno
    if 'salidas' in modulos:
        story.append(Paragraph("Resumen de Salidas a Terreno", styles['HeaderStyle']))
        salidas = SalidaTerreno.objects.all()
        salidas_data = [['Motivo', 'Fecha Salida', 'A Cargo', 'Km Recorridos']]
        for s in salidas:
            salidas_data.append([
                s.motivo, s.fecha_hora_salida.strftime('%d/%m/%Y'), s.personal_a_cargo.get_full_name() if s.personal_a_cargo else 'N/A', f"{s.kilometros_recorridos or 0} km"
            ])
        story.append(create_table(salidas_data))
        story.append(Spacer(1, 0.3 * inch))

    # 4. Mantenimientos
    if 'mantenimientos' in modulos:
        story.append(Paragraph("Resumen de Mantenimientos", styles['HeaderStyle']))
        mantenimientos = Mantenimiento.objects.all()
        mantenimientos_data = [['Vehículo', 'Fecha', 'Tipo', 'Costo']]
        for m in mantenimientos:
            mantenimientos_data.append([
                m.vehiculo, m.fecha.strftime('%d/%m/%Y'), m.tipo, f"${m.costo or 0:,.0f}"
            ])
        story.append(create_table(mantenimientos_data))
        story.append(Spacer(1, 0.3 * inch))

    # 5. Inventario
    if 'inventario' in modulos:
        story.append(Paragraph("Resumen de Inventario", styles['HeaderStyle']))
        inventario = Inventario.objects.all()
        # Agrupamos los items por nombre, ubicación y estado para obtener la cantidad
        inventario_agrupado = Inventario.objects.values('nombre', 'ubicacion', 'estado').annotate(cantidad=Count('id')).order_by('nombre')
        
        inventario_data = [['Nombre', 'Cantidad', 'Ubicación', 'Estado']]
        for i in inventario_agrupado:
            inventario_data.append([Inventario.ITEM_CHOICES_DICT.get(i['nombre'], i['nombre']), i['cantidad'], i['ubicacion'], i['estado']])
        story.append(create_table(inventario_data))
        story.append(Spacer(1, 0.3 * inch))

    # 6. Caja Chica
    if 'caja_chica' in modulos:
        story.append(Paragraph("Resumen de Caja Chica", styles['HeaderStyle']))
        caja_chica = CajaChica.objects.all()
        if not request.user.is_superuser:
            if request.user.compania:
                caja_chica = caja_chica.filter(compania=request.user.compania)
            else:
                caja_chica = CajaChica.objects.none()
                
        caja_chica_data = [['Fecha', 'Tipo', 'Descripción', 'Monto']]
        for c in caja_chica:
            monto_str = f"+${c.monto:,.0f}" if c.tipo == 'Ingreso' else f"-${c.monto:,.0f}"
            caja_chica_data.append([c.fecha.strftime('%d/%m/%Y'), c.tipo, c.descripcion, monto_str])
        story.append(create_table(caja_chica_data))
        story.append(Spacer(1, 0.3 * inch))

    # 7. Cursos
    if 'cursos' in modulos:
        story.append(Paragraph("Resumen de Cursos", styles['HeaderStyle']))
        capacitaciones = Capacitacion.objects.all()
        cursos_data = [['Nombre', 'Lugar', 'Fecha Inicio', 'Asistentes']]
        for c in capacitaciones:
            cursos_data.append([c.nombre, c.lugar, c.fecha_inicio.strftime('%d/%m/%Y'), str(c.asistentes.count())])
        if len(cursos_data) > 1: story.append(create_table(cursos_data))
        story.append(Spacer(1, 0.3 * inch))

    # 8. Documentos
    if 'documentos' in modulos:
        story.append(Paragraph("Resumen de Documentos", styles['HeaderStyle']))
        documentos = Documento.objects.all()
        doc_data = [['Nombre', 'Subido por', 'Fecha', 'Tamaño']]
        for d in documentos:
            doc_data.append([d.nombre, d.subido_por.get_full_name() if d.subido_por else 'N/A', d.subido_en.strftime('%d/%m/%Y'), f"{d.tamano_seguro / 1024:.1f} KB"])
        if len(doc_data) > 1: story.append(create_table(doc_data))
        story.append(Spacer(1, 0.3 * inch))

    # --- Generar el PDF ---
    doc.build(story)

    return response

@login_required
def generar_reporte_anual_excel(request):
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permiso para generar este tipo de reportes.')
        return redirect('dashboard_summary')

    # --- Módulos a exportar ---
    modulos_str = request.GET.get('modulos', 'documentos,inventario,salidas,emergencias,cursos,mantenimientos,proyectos,caja_chica')
    modulos = [m.strip() for m in modulos_str.split(',') if m.strip()]

    # --- Configuración del Documento Excel ---
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="reporte_anual_{timezone.now().year}.xlsx"'

    workbook = openpyxl.Workbook()

    # --- Función para añadir datos a una hoja ---
    def add_sheet(title, headers, data_rows):
        sheet = workbook.create_sheet(title=title)
        sheet.append(headers)
        for row in data_rows:
            sheet.append(row)
        # (Opcional) Añadir estilos, autoajuste de columnas, etc.

    # 1. Proyectos
    if 'proyectos' in modulos:
        proyectos = Proyecto.objects.all()
        proyectos_data = []
        for p in proyectos:
            proyectos_data.append([
                p.nombre, p.tipo, p.responsable.get_full_name() if p.responsable else 'N/A', p.presupuesto or 0
            ])
        add_sheet("Proyectos", ['Nombre', 'Tipo', 'Responsable', 'Presupuesto'], proyectos_data)

    # 2. Emergencias
    if 'emergencias' in modulos:
        emergencias = Emergencia.objects.all()
        emergencias_data = []
        for e in emergencias:
            emergencias_data.append([
                e.get_tipo_display(), e.direccion, e.fecha_hora_alarma.strftime('%d/%m/%Y %H:%M'), e.estado
            ])
        add_sheet("Emergencias", ['Tipo', 'Dirección', 'Fecha', 'Estado'], emergencias_data)

    # 3. Salidas a Terreno
    if 'salidas' in modulos:
        salidas = SalidaTerreno.objects.all()
        salidas_data = []
        for s in salidas:
            salidas_data.append([
                s.motivo, s.fecha_hora_salida.strftime('%d/%m/%Y %H:%M'), s.personal_a_cargo.get_full_name() if s.personal_a_cargo else 'N/A', s.kilometros_recorridos or 0
            ])
        add_sheet("Salidas a Terreno", ['Motivo', 'Fecha Salida', 'A Cargo', 'Km Recorridos'], salidas_data)

    # 4. Mantenimientos
    if 'mantenimientos' in modulos:
        mantenimientos = Mantenimiento.objects.all()
        mantenimientos_data = []
        for m in mantenimientos:
            mantenimientos_data.append([
                m.vehiculo, m.fecha.strftime('%d/%m/%Y'), m.tipo, m.costo or 0
            ])
        add_sheet("Mantenimientos", ['Vehículo', 'Fecha', 'Tipo', 'Costo'], mantenimientos_data)

    # 5. Inventario
    if 'inventario' in modulos:
        inventario_agrupado = Inventario.objects.values('nombre', 'ubicacion', 'estado').annotate(cantidad=Count('id')).order_by('nombre')
        inventario_data = []
        for i in inventario_agrupado:
            inventario_data.append([Inventario.ITEM_CHOICES_DICT.get(i['nombre'], i['nombre']), i['cantidad'], i['ubicacion'], i['estado']])
        add_sheet("Inventario", ['Nombre', 'Cantidad', 'Ubicación', 'Estado'], inventario_data)

    # 6. Caja Chica
    if 'caja_chica' in modulos:
        caja_chica = CajaChica.objects.all()
        if not request.user.is_superuser:
            if request.user.compania:
                caja_chica = caja_chica.filter(compania=request.user.compania)
            else:
                caja_chica = CajaChica.objects.none()
                
        caja_chica_data = []
        for c in caja_chica:
            monto = c.monto if c.tipo == 'Ingreso' else -c.monto
            caja_chica_data.append([c.fecha.strftime('%d/%m/%Y'), c.tipo, c.descripcion, monto])
        add_sheet("Caja Chica", ['Fecha', 'Tipo', 'Descripción', 'Monto'], caja_chica_data)

    # 7. Cursos
    if 'cursos' in modulos:
        capacitaciones = Capacitacion.objects.all()
        cursos_data = []
        for c in capacitaciones:
            cursos_data.append([c.nombre, c.lugar, c.fecha_inicio.strftime('%d/%m/%Y %H:%M'), c.asistentes.count()])
        add_sheet("Cursos", ['Nombre', 'Lugar', 'Fecha Inicio', 'Asistentes'], cursos_data)

    # 8. Documentos
    if 'documentos' in modulos:
        documentos = Documento.objects.all()
        doc_data = []
        for d in documentos:
            doc_data.append([d.nombre, d.subido_por.get_full_name() if d.subido_por else 'N/A', d.subido_en.strftime('%d/%m/%Y %H:%M'), round(d.tamano_seguro / 1024, 2)])
        add_sheet("Documentos", ['Nombre', 'Subido por', 'Fecha', 'Tamaño (KB)'], doc_data)

    # Eliminar la hoja por defecto que crea openpyxl
    if len(workbook.sheetnames) > 1 and 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    # --- Guardar el libro de trabajo en la respuesta HTTP ---
    workbook.save(response)

    return response

@login_required
def proyectos_view(request):
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.ver_proyectos):
        messages.error(request, 'No tienes permiso para acceder a Proyectos.')
        return redirect('dashboard')

    if request.method == 'POST':
        if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_proyectos):
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': {'__all__': ['No tienes permiso para crear proyectos.']}}, status=403)
            messages.error(request, 'No tienes permiso para crear proyectos.')
            return redirect('proyectos')
            
        proyecto_form = ProyectoForm(request.POST) 
        archivo_form = ArchivoProyectoForm(request.POST, request.FILES) # Este formulario no tiene validación crítica, pero lo mantenemos por consistencia
        if proyecto_form.is_valid():
            proyecto = proyecto_form.save(commit=False)
            proyecto.creado_por = request.user
            proyecto.save()
            
            # Correctamente obtenemos la lista de archivos del campo 'archivos'
            files = request.FILES.getlist('archivos') 
            for f in files:
                ArchivoProyecto.objects.create(proyecto=proyecto, archivo=f)
            
            # Si la solicitud es AJAX, devolvemos JSON
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'proyecto': {
                        'id': proyecto.id,
                        'nombre': proyecto.nombre,
                        'descripcion': proyecto.descripcion,
                        'tipo': proyecto.tipo,
                        'fecha_inicio': proyecto.fecha_inicio.isoformat() if proyecto.fecha_inicio else None,
                        'fecha_fin': proyecto.fecha_fin.isoformat() if proyecto.fecha_fin else None,
                        'responsable_nombre': proyecto.responsable.get_full_name() if proyecto.responsable else "No asignado",
                        'responsable_id': proyecto.responsable.id if proyecto.responsable else None,
                        'presupuesto': f"{proyecto.presupuesto:,.0f}" if proyecto.presupuesto else None,
                        'presupuesto_raw': str(proyecto.presupuesto) if proyecto.presupuesto else '',
                        'prioridad': proyecto.prioridad,
                        'archivos': list(proyecto.archivos.all().values('archivo', 'subido_en')),
                    }
                })
            else: # Comportamiento normal
                messages.success(request, f'Proyecto "{proyecto.nombre}" creado exitosamente.')
                return redirect('proyectos')
        else:
            # Si la solicitud es AJAX y hay errores, devolvemos los errores en JSON
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': proyecto_form.errors}, status=400)
            else: # Comportamiento normal
                messages.error(request, 'Hubo un error al crear el proyecto. Por favor, revisa el formulario.')

    else:
        proyecto_form = ProyectoForm()
        archivo_form = ArchivoProyectoForm()

    # Extraemos todos los proyectos para quienes tienen permiso de entrar a esta vista
    proyectos = Proyecto.objects.all().order_by('-creado_en')
    
    context = {
        'proyecto_form': proyecto_form, 
        'archivo_form': archivo_form, 
        'proyectos': proyectos,
        'today': timezone.now().date() # Pasamos la fecha de hoy al contexto
    }
    return render(request, 'usuarios/proyectos.html', context)

@login_required
def proyecto_edit_view(request, proyecto_id):
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    # Permiso de edición
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_proyectos):
        messages.error(request, 'No tienes permiso para editar este proyecto.')
        return redirect('proyectos')

    if request.method == 'POST':
        form = ProyectoForm(request.POST, instance=proyecto)
        if form.is_valid():
            proyecto_actualizado = form.save()
            
            # Añadir nuevos archivos si se subieron
            files = request.FILES.getlist('archivos')
            for f in files:
                ArchivoProyecto.objects.create(proyecto=proyecto_actualizado, archivo=f)

            messages.success(request, f'Proyecto "{proyecto_actualizado.nombre}" actualizado exitosamente.')
        else:
            messages.error(request, 'Error al actualizar el proyecto. Revisa el formulario.')
    return redirect('proyectos')

@login_required
def proyecto_delete_view(request, proyecto_id):
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_proyectos):
        messages.error(request, 'No tienes permiso para eliminar este proyecto.')
        return redirect('proyectos')

    if request.method == 'POST':
        proyecto_nombre = proyecto.nombre
        proyecto.delete() # Los archivos se eliminan en cascada gracias a la señal
        messages.success(request, f'Proyecto "{proyecto_nombre}" y todos sus archivos han sido eliminados.')
    return redirect('proyectos')

@login_required
def documentos_view(request):
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.ver_documentacion):
        messages.error(request, 'No tienes permiso para acceder a Documentos.')
        return redirect('dashboard')

    if request.method == 'POST':
        if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_documentacion):
            messages.error(request, 'No tienes permiso para modificar documentos.')
            return redirect('documentos')
            
        accion = request.POST.get('action')
        if accion == 'crear_texto':
            nombre = request.POST.get('nombre_texto')
            descripcion = request.POST.get('descripcion_texto', '')
            contenido = request.POST.get('contenido_texto', '')
            compania_id = request.POST.get('compania')

            if not nombre or not contenido:
                messages.error(request, 'El título y el contenido son obligatorios.')
                return redirect('documentos')

            buffer = BytesIO()
            doc_pdf = SimpleDocTemplate(buffer, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []
            
            story.append(Paragraph(nombre, styles['Title']))
            story.append(Spacer(1, 12))
            
            for line in contenido.split('\n'):
                if line.strip():
                    clean_line = line.strip().replace('<', '&lt;').replace('>', '&gt;')
                    story.append(Paragraph(clean_line, styles['Normal']))
                else:
                    story.append(Spacer(1, 12))
                    
            doc_pdf.build(story)
            pdf_bytes = buffer.getvalue()
            buffer.close()

            safe_filename = re.sub(r'[^a-zA-Z0-9_\-]', '_', nombre) + '.pdf'

            documento = Documento(
                nombre=nombre,
                descripcion=descripcion,
                subido_por=request.user
            )
            
            if request.user.is_superuser and compania_id:
                try:
                    documento.compania = Compania.objects.get(id=compania_id)
                except Compania.DoesNotExist:
                    pass
            elif not request.user.is_superuser and request.user.compania:
                documento.compania = request.user.compania
                
            documento.archivo.save(safe_filename, ContentFile(pdf_bytes), save=False)
            documento.save()
            
            messages.success(request, f'Documento "{nombre}" redactado y guardado como PDF exitosamente.')
            return redirect('documentos')
        else:
            form = DocumentoForm(request.POST, request.FILES, user=request.user)
            if form.is_valid():
                documento = form.save(commit=False)
                documento.subido_por = request.user
                
                if not request.user.is_superuser and request.user.compania:
                    documento.compania = request.user.compania
                    
                documento.save()
                messages.success(request, f'Documento "{documento.nombre}" subido exitosamente.')
                return redirect('documentos')
            else:
                messages.error(request, 'Error al subir el documento. Revisa el formulario.')
    else:
        form = DocumentoForm(user=request.user)
    
    # --- Lógica de Búsqueda y Filtros ---
    query = request.GET.get('q')
    tipo_archivo = request.GET.get('tipo_archivo')
    usuario_id = request.GET.get('usuario')
    ordenar_por = request.GET.get('ordenar_por', '-subido_en') # Por defecto, los más nuevos primero

    documentos = Documento.objects.select_related('subido_por', 'compania').all()

    if not request.user.is_superuser:
        if request.user.compania:
            # Mostrar los de su compañía o los que son generales (compania=None)
            documentos = documentos.filter(Q(compania=request.user.compania) | Q(compania__isnull=True))
        else:
            documentos = documentos.filter(compania__isnull=True)

    # Filtrar por término de búsqueda
    if query:
        documentos = documentos.filter(
            Q(nombre__icontains=query) |
            Q(descripcion__icontains=query)
        )

    # Filtrar por tipo de archivo
    if tipo_archivo:
        if tipo_archivo == 'pdf':
            documentos = documentos.filter(archivo__iendswith='.pdf')
        elif tipo_archivo == 'word':
            documentos = documentos.filter(Q(archivo__iendswith='.doc') | Q(archivo__iendswith='.docx'))
        elif tipo_archivo == 'excel':
            documentos = documentos.filter(Q(archivo__iendswith='.xls') | Q(archivo__iendswith='.xlsx'))
        elif tipo_archivo == 'imagen':
            documentos = documentos.filter(Q(archivo__iendswith='.jpg') | Q(archivo__iendswith='.jpeg') | Q(archivo__iendswith='.png'))

    # Filtrar por usuario
    if usuario_id:
        documentos = documentos.filter(subido_por__id=usuario_id)

    # Aplicar ordenamiento
    valid_sorts = ['-subido_en', 'subido_en', 'nombre', '-nombre']
    if ordenar_por in valid_sorts:
        documentos = documentos.order_by(ordenar_por)

    if not request.user.is_superuser and request.user.compania:
        usuarios_con_docs = Usuario.objects.filter(Q(documento__compania=request.user.compania) | Q(documento__compania__isnull=True), documento__isnull=False).distinct().order_by('nombre')
    else:
        usuarios_con_docs = Usuario.objects.filter(documento__isnull=False).distinct().order_by('nombre')

    context = {'form': form, 'documentos': documentos, 'usuarios_con_docs': usuarios_con_docs}
    return render(request, 'usuarios/documentos.html', context)

@login_required
def documento_delete_view(request, doc_id):
    documento = get_object_or_404(Documento, id=doc_id)
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_documentacion):
        messages.error(request, 'No tienes permiso para eliminar este documento.')
        return redirect('documentos')

    if request.method == 'POST':
        doc_name = documento.nombre
        documento.delete()
        messages.success(request, f'Documento "{doc_name}" eliminado exitosamente.')
    return redirect('documentos')

@login_required
def inventario_view(request):
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.ver_inventario):
        messages.error(request, 'No tienes permiso para acceder al Inventario.')
        return redirect('dashboard')

    # Esta vista ahora solo maneja la creación de nuevos ítems.
    if request.method == 'POST':
        if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_inventario):
            messages.error(request, 'No tienes permiso para agregar ítems.')
            return redirect('inventario')
            
        form = InventarioForm(request.POST, user=request.user)
        if form.is_valid():
            cantidad = form.cleaned_data.get('cantidad', 1)
            
            for _ in range(cantidad):
                item = Inventario(
                    nombre=form.cleaned_data['nombre'],
                    ubicacion=form.cleaned_data['ubicacion'],
                    estado=form.cleaned_data['estado'],
                    fecha_adquisicion=form.cleaned_data.get('fecha_adquisicion'),
                    agregado_por=request.user
                )
                # Asignación Automática de Compañía
                if request.user.is_superuser:
                    item.compania = form.cleaned_data.get('compania')
                else:
                    item.compania = request.user.compania
                    
                item.save() # El método save() del modelo se encarga del QR

            messages.success(request, f'{cantidad} ítem(s) de "{item.get_nombre_display()}" agregados exitosamente.')
            return redirect('inventario')
        else:
            messages.error(request, 'Error al agregar el ítem. Por favor, revisa el formulario.')

    # --- Lógica de visualización ---
    inventario_por_compania = {}
    companias_a_mostrar = Compania.objects.all()

    # Si el usuario no es superusuario, solo mostramos su compañía
    if not request.user.is_superuser and request.user.compania:
        companias_a_mostrar = Compania.objects.filter(pk=request.user.compania.pk)

    # Búsqueda general
    query = request.GET.get('q')
    compania_activa_id = request.GET.get('compania_activa')

    # Si se está buscando en una compañía específica, filtramos la lista de compañías a mostrar
    if query and compania_activa_id:
        try:
            companias_a_mostrar = companias_a_mostrar.filter(id=int(compania_activa_id))
        except (ValueError, TypeError):
            pass # Si el ID no es válido, simplemente mostramos todo como antes

    for compania in companias_a_mostrar:
        base_queryset = Inventario.objects.filter(compania=compania).select_related('asignado_a', 'compania')

        if query:
            base_queryset = base_queryset.filter(
                Q(nombre__icontains=query) |
                Q(asignado_a__nombre__icontains=query) |
                Q(asignado_a__apellido__icontains=query) |
                Q(ubicacion__icontains=query)
            )

        items_asignados_qs = base_queryset.filter(asignado_a__isnull=False).order_by('asignado_a__nombre', 'asignado_a__apellido', 'nombre')
        
        user_groups = {}
        for item in items_asignados_qs:
            if item.asignado_a not in user_groups:
                user_groups[item.asignado_a] = []
            user_groups[item.asignado_a].append(item)
            
        items_asignados = [{'usuario': u, 'items': items} for u, items in user_groups.items()]
        
        # Procesamos los ítems disponibles para añadir el nombre completo
        items_disponibles_agrupados_qs = base_queryset.filter(asignado_a__isnull=True).values('nombre', 'ubicacion', 'estado').annotate(cantidad=Count('id')).order_by('nombre', 'ubicacion', 'estado')
        items_disponibles_agrupados = []
        for item in items_disponibles_agrupados_qs:
            item['nombre_completo'] = Inventario.get_nombre_completo(item['nombre'])
            items_disponibles_agrupados.append(item)

        inventario_por_compania[compania] = {
            'asignados': items_asignados,
            'disponibles_agrupados': items_disponibles_agrupados
        }

    # Datos para el nuevo panel de asignación
    usuarios_activos = Usuario.objects.filter(is_active=True).order_by('nombre')
    # Si el usuario no es superusuario, filtramos para que solo vea/asigne a miembros de su compañía
    if not request.user.is_superuser and request.user.compania:
        usuarios_activos = usuarios_activos.filter(compania=request.user.compania)
    
    # CORRECCIÓN: Agrupar los ítems disponibles por nombre Y por compañía para el selector de asignación.
    # Esto asegura que el atributo data-compania en el HTML tenga el valor correcto.
    items_disponibles_qs = Inventario.objects.filter(asignado_a__isnull=True).values('nombre', 'compania_id').annotate(cantidad=Count('id')).order_by('nombre')
    
    items_disponibles_para_asignar = []
    for item in items_disponibles_qs:
        item['nombre_completo'] = Inventario.get_nombre_completo(item['nombre'])
        # El 'compania_id' ya está incluido gracias a la consulta corregida.
        items_disponibles_para_asignar.append(item)

    context = {
        'form': InventarioForm(user=request.user), 
        'edit_form': InventarioEditForm(user=request.user), # Para el modal de edición
        'group_edit_form': InventarioGroupEditForm(), # Para el modal de edición de grupo
        'inventario_por_compania': inventario_por_compania,
        'items_disponibles_para_asignar': items_disponibles_para_asignar,
        'usuarios_activos': usuarios_activos,
        'first_compania_id': list(inventario_por_compania.keys())[0].id if inventario_por_compania else None,
    }
    return render(request, 'usuarios/inventario.html', context)

@login_required
def inventario_edit_view(request, item_id):
    item = get_object_or_404(Inventario, id=item_id)
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_inventario):
        messages.error(request, 'No tienes permiso para editar este ítem.')
        return redirect('inventario')

    if request.method == 'POST':
        # Usamos un formulario sin el campo 'cantidad' para la edición
        form = InventarioEditForm(request.POST, instance=item, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, f'Ítem "{item.get_nombre_display()}" actualizado exitosamente.')
        else:
            messages.error(request, 'Error al actualizar el ítem. Por favor, revisa el formulario.')
    return redirect('inventario')

@login_required
def inventario_delete_view(request, item_id):
    item = get_object_or_404(Inventario, id=item_id)
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_inventario):
        messages.error(request, 'No tienes permiso para eliminar este ítem.')
        return redirect('inventario')

    if request.method == 'POST':
        item_name = item.get_nombre_display()
        item.delete()
        messages.success(request, f'Ítem "{item_name}" eliminado exitosamente.')
    return redirect('inventario')

@login_required
def inventario_unassign_view(request, item_id):
    item = get_object_or_404(Inventario, id=item_id)
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_inventario):
        messages.error(request, 'No tienes permiso para modificar este ítem.')
        return redirect('inventario')

    if request.method == 'POST':
        usuario_anterior = item.asignado_a
        item_name = item.get_nombre_display()
        
        if usuario_anterior:
            item.asignado_a = None
            item.save() # El save() limpiará también el QR asignado a este ítem automáticamente
            Notificacion.objects.create(
                usuario=usuario_anterior,
                mensaje=f"Has devuelto a bodega el equipo: {item_name}.",
                link="/inventario/"
            )
            messages.success(request, f'Ítem "{item_name}" devuelto a bodega exitosamente (desasignado de {usuario_anterior.get_full_name()}).')
        else:
            messages.warning(request, f'El ítem "{item_name}" ya se encontraba en bodega.')
            
    return redirect('inventario')

@login_required
def inventario_assign_view(request):
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_inventario):
        messages.error(request, 'No tienes permiso para realizar esta acción.')
        return redirect('inventario')

    if request.method == 'POST':
        user_id = request.POST.get('usuario_a_asignar')
        items_nombres = request.POST.getlist('item_a_asignar')
        cantidades = request.POST.getlist('cantidad_a_asignar')

        if not user_id or not items_nombres or not cantidades:
            messages.error(request, 'Debes seleccionar un usuario y al menos un ítem válido para asignar.')
            return redirect('inventario')

        usuario = get_object_or_404(Usuario, id=user_id)
        
        # Validación extra: Asegurarnos de que el usuario tenga una compañía asignada
        if not usuario.compania:
            messages.error(request, f'El usuario {usuario.get_full_name()} no tiene una compañía asignada y no puede recibir ítems.')
            return redirect('inventario')

        num_asignados_total = 0
        items_asignados_detalle = []
        errores = []

        for item_nombre, cantidad_str in zip(items_nombres, cantidades):
            try:
                cantidad_a_asignar = int(cantidad_str)
            except (ValueError, TypeError):
                cantidad_a_asignar = 0

            if cantidad_a_asignar <= 0:
                continue

            items_para_asignar = Inventario.objects.filter(
                nombre=item_nombre, 
                asignado_a__isnull=True,
                compania=usuario.compania
            )[:cantidad_a_asignar]

            if not items_para_asignar:
                errores.append(f'No hay ítems de "{Inventario.ITEM_CHOICES_DICT.get(item_nombre, item_nombre)}" disponibles en la {usuario.compania.nombre}.')
                continue

            if len(items_para_asignar) < cantidad_a_asignar:
                errores.append(f'No había suficientes "{Inventario.ITEM_CHOICES_DICT.get(item_nombre, item_nombre)}". Solo se asignaron {len(items_para_asignar)} de {cantidad_a_asignar}.')

            asignados_de_este_tipo = 0
            for item in items_para_asignar:
                item.asignado_a = usuario
                item.save()
                num_asignados_total += 1
                asignados_de_este_tipo += 1
                
            if asignados_de_este_tipo > 0:
                items_asignados_detalle.append(f"{asignados_de_este_tipo}x {Inventario.ITEM_CHOICES_DICT.get(item_nombre, item_nombre)}")

        if errores:
            for error in errores:
                messages.warning(request, error)

        if num_asignados_total > 0:
            detalle_msg = ", ".join(items_asignados_detalle)
            Notificacion.objects.create(
                usuario=usuario,
                mensaje=f"Se te han asignado nuevos equipos: {detalle_msg}.",
                link="/inventario/"
            )
            messages.success(request, f'Se asignaron exitosamente a {usuario.get_full_name()}: {detalle_msg}.')
        elif not errores:
            messages.error(request, 'No se pudo asignar ningún ítem.')
            
    return redirect('inventario')

@login_required
def inventario_group_edit_view(request):
    if request.method != 'POST' or (not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_inventario)):
        messages.error(request, 'Acción no permitida.')
        return redirect('inventario')

    form = InventarioGroupEditForm(request.POST)
    if form.is_valid():
        nombre = request.POST.get('group_nombre')
        ubicacion_actual = request.POST.get('group_ubicacion')
        estado_actual = request.POST.get('group_estado')

        items_a_actualizar = Inventario.objects.filter(
            asignado_a__isnull=True,
            nombre=nombre,
            ubicacion=ubicacion_actual,
            estado=estado_actual
        )
        
        count = items_a_actualizar.update(
            ubicacion=form.cleaned_data['ubicacion'],
            estado=form.cleaned_data['estado']
        )
        messages.success(request, f'{count} ítems del grupo "{nombre}" han sido actualizados.')
    else:
        messages.error(request, 'Error al actualizar el grupo de ítems. Revisa el formulario.')
    
    return redirect('inventario')

@login_required
def inventario_group_delete_view(request):
    if request.method != 'POST' or (not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_inventario)):
        messages.error(request, 'Acción no permitida.')
        return redirect('inventario')

    Inventario.objects.filter(asignado_a__isnull=True, nombre=request.POST.get('group_nombre'), ubicacion=request.POST.get('group_ubicacion'), estado=request.POST.get('group_estado')).delete()
    messages.success(request, f'El grupo de ítems "{request.POST.get("group_nombre")}" ha sido eliminado.')
    return redirect('inventario')

@login_required
def scan_qr_page_view(request):
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.ver_escanear_qr):
        messages.error(request, 'No tienes permiso para usar el Escáner QR.')
        return redirect('dashboard')

    return render(request, 'usuarios/scan_qr.html')

@login_required
def inventario_qr_lookup_view(request):
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.ver_escanear_qr):
        return JsonResponse({'error': 'No tienes permisos para realizar esta acción.'}, status=403)

    qr_content = request.GET.get('qr_content', '')
    item_id = None
    
    # Primero, intentamos extraer el ID del formato "BOMBEROS-ITEM-ID:XX"
    match = re.search(r'BOMBEROS-ITEM-ID:(\d+)', qr_content)
    if match:
        item_id = match.group(1)
    # Si no, asumimos que el contenido del QR es directamente el ID (formato nuevo y más robusto)
    elif qr_content.isdigit():
        item_id = qr_content

    if not item_id:
        return JsonResponse({'error': 'Código QR no válido o no reconocido.'}, status=400)

    try:
        item = Inventario.objects.select_related('asignado_a').get(id=item_id)
        if not item.asignado_a:
            return JsonResponse({'error': 'Este ítem está en bodega y no está asignado a ningún usuario.'}, status=404)

        data = {
            'item_nombre': item.get_nombre_display(),
            'usuario_nombre': item.asignado_a.get_full_name(),
            'usuario_foto_url': item.asignado_a.foto_perfil.url if item.asignado_a.foto_perfil else None,
            'usuario_iniciales': f"{item.asignado_a.nombre[0] if item.asignado_a.nombre else ''}{item.asignado_a.apellido[0] if item.asignado_a.apellido else ''}".upper(),
        }
        return JsonResponse(data)
    except Inventario.DoesNotExist:
        return JsonResponse({'error': 'Ítem de inventario no encontrado.'}, status=404)

@login_required
def salidas_terreno_view(request):
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.ver_salidas_terreno):
        messages.error(request, 'No tienes permiso para acceder a Salidas a Terreno.')
        return redirect('dashboard')

    # La vista principal ahora también maneja los errores del formulario de creación
    if request.method == 'POST':
        if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_salidas_terreno):
            messages.error(request, 'No tienes permiso para registrar salidas.')
            return redirect('salidas_terreno')
            
        form = SalidaTerrenoForm(request.POST)
        if form.is_valid():
            salida = form.save(commit=False)
            salida.creado_por = request.user
            salida.save()
            messages.success(request, f'Salida a terreno por "{salida.motivo}" registrada.')
            return redirect('salidas_terreno')
        else:
            messages.error(request, 'Error al registrar la salida. Por favor, revisa el formulario.')
    else:
        form = SalidaTerrenoForm()

    # --- Lógica de Búsqueda y Filtros ---
    query = request.GET.get('q')
    usuario_id = request.GET.get('usuario')
    ordenar_por = request.GET.get('ordenar_por', '-fecha_hora_salida')

    salidas = SalidaTerreno.objects.select_related('personal_a_cargo').all()

    if query:
        salidas = salidas.filter(
            Q(motivo__icontains=query) |
            Q(direccion__icontains=query) |
            Q(unidades_involucradas__icontains=query) |
            Q(descripcion__icontains=query) |
            Q(personal_a_cargo__nombre__icontains=query) |
            Q(personal_a_cargo__apellido__icontains=query)
        )

    if usuario_id:
        salidas = salidas.filter(personal_a_cargo__id=usuario_id)

    valid_sorts = ['-fecha_hora_salida', 'fecha_hora_salida', 'motivo', '-motivo']
    if ordenar_por in valid_sorts:
        salidas = salidas.order_by(ordenar_por)

    usuarios_con_salidas = Usuario.objects.filter(salidas_a_cargo__isnull=False).distinct().order_by('nombre')
    context = {
        'form': form, 
        'salidas': salidas,
        'usuarios_con_salidas': usuarios_con_salidas
    }
    return render(request, 'usuarios/salidas_terreno.html', context)

@login_required
def salida_terreno_edit_view(request, salida_id):
    salida = get_object_or_404(SalidaTerreno, id=salida_id)
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_salidas_terreno):
        messages.error(request, 'No tienes permiso para editar esta salida.')
        return redirect('salidas_terreno')

    if request.method == 'POST':
        form = SalidaTerrenoForm(request.POST, instance=salida)
        if form.is_valid():
            form.save()
            messages.success(request, f'Salida por "{salida.motivo}" actualizada exitosamente.')
            return redirect('salidas_terreno')
        else:
            messages.error(request, 'Error al actualizar la salida. Revisa el formulario.')
    return redirect('salidas_terreno')

@login_required
def salida_terreno_delete_view(request, salida_id):
    salida = get_object_or_404(SalidaTerreno, id=salida_id)
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_salidas_terreno):
        messages.error(request, 'No tienes permiso para eliminar esta salida.')
        return redirect('salidas_terreno')

    if request.method == 'POST':
        salida_motivo = salida.motivo
        salida.delete()
        messages.success(request, f'Salida por "{salida_motivo}" eliminada exitosamente.')
    return redirect('salidas_terreno')

@login_required
def emergencias_view(request):
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.ver_emergencias):
        messages.error(request, 'No tienes permiso para acceder a Emergencias.')
        return redirect('dashboard')

    if request.method == 'POST':
        if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_emergencias):
            messages.error(request, 'No tienes permiso para registrar emergencias.')
            return redirect('emergencias')
            
        form = EmergenciaForm(request.POST)
        if form.is_valid():
            emergencia = form.save(commit=False)
            emergencia.creado_por = request.user
            emergencia.save()
            messages.success(request, f'Emergencia "{emergencia.get_tipo_display()}" registrada exitosamente.')
            return redirect('emergencias')
        else:
            messages.error(request, 'Error al registrar la emergencia. Por favor, revisa el formulario.')
    else:
        form = EmergenciaForm()

    # --- Lógica de Búsqueda y Filtros ---
    query = request.GET.get('q')
    tipo_emergencia = request.GET.get('tipo')
    estado_emergencia = request.GET.get('estado')
    usuario_id = request.GET.get('usuario')
    ordenar_por = request.GET.get('ordenar_por', '-fecha_hora_alarma')

    emergencias = Emergencia.objects.select_related('oficial_a_cargo').all()

    if query:
        emergencias = emergencias.filter(
            Q(direccion__icontains=query) |
            Q(descripcion__icontains=query) |
            Q(unidades_despachadas__icontains=query)
        )
    
    if tipo_emergencia:
        emergencias = emergencias.filter(tipo=tipo_emergencia)
    
    if estado_emergencia:
        emergencias = emergencias.filter(estado=estado_emergencia)

    if usuario_id:
        emergencias = emergencias.filter(oficial_a_cargo__id=usuario_id)

    valid_sorts = ['-fecha_hora_alarma', 'fecha_hora_alarma', 'tipo', '-tipo']
    if ordenar_por in valid_sorts:
        emergencias = emergencias.order_by(ordenar_por)

    usuarios_con_emergencias = Usuario.objects.filter(emergencias_a_cargo__isnull=False).distinct().order_by('nombre')

    context = {'form': form, 'emergencias': emergencias, 'usuarios_con_emergencias': usuarios_con_emergencias}
    return render(request, 'usuarios/emergencias.html', context)

@login_required
def emergencia_edit_view(request, emergencia_id):
    emergencia = get_object_or_404(Emergencia, id=emergencia_id)
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_emergencias):
        messages.error(request, 'No tienes permiso para editar esta emergencia.')
        return redirect('emergencias')

    if request.method == 'POST':
        form = EmergenciaForm(request.POST, instance=emergencia)
        if form.is_valid():
            form.save()
            messages.success(request, f'Emergencia "{emergencia.get_tipo_display()}" actualizada exitosamente.')
        else:
            messages.error(request, 'Error al actualizar la emergencia. Revisa el formulario.')
    return redirect('emergencias')

@login_required
def emergencia_delete_view(request, emergencia_id):
    emergencia = get_object_or_404(Emergencia, id=emergencia_id)
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_emergencias):
        messages.error(request, 'No tienes permiso para eliminar esta emergencia.')
        return redirect('emergencias')

    if request.method == 'POST':
        emergencia_tipo = emergencia.get_tipo_display()
        emergencia.delete()
        messages.success(request, f'Emergencia tipo "{emergencia_tipo}" eliminada exitosamente.')
    return redirect('emergencias')

@login_required
def capacitaciones_view(request):
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.ver_capacitaciones):
        messages.error(request, 'No tienes permiso para acceder a Cursos.')
        return redirect('dashboard')

    if request.method == 'POST':
        if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_capacitaciones):
            messages.error(request, 'No tienes permiso para registrar cursos.')
            return redirect('capacitaciones')
            
        form = CapacitacionForm(request.POST)
        if form.is_valid():
            capacitacion = form.save(commit=False)
            capacitacion.creado_por = request.user
            capacitacion.save()
            form.save_m2m() # Necesario para guardar las relaciones ManyToMany
            
            # Lógica de Notificaciones / Invitaciones
            enviar_invitacion = request.POST.get('enviar_invitacion') == 'on'
            audiencia = request.POST.get('audiencia', 'general')

            if enviar_invitacion:
                if audiencia == 'general':
                    if capacitacion.companias_invitadas.exists():
                        usuarios_a_notificar = Usuario.objects.filter(is_active=True, compania__in=capacitacion.companias_invitadas.all())
                    else:
                        usuarios_a_notificar = Usuario.objects.none()
                else:
                    usuarios_a_notificar = capacitacion.asistentes.all()

                notificaciones = []
                for u in usuarios_a_notificar:
                    mensaje = f"Has sido invitado al curso: {capacitacion.nombre}."
                    if capacitacion.cupos:
                        mensaje += f" ¡Cupos limitados ({capacitacion.cupos})!"
                    
                    notificaciones.append(Notificacion(usuario=u, mensaje=mensaje, link=f"/capacitaciones/"))
                
                if notificaciones:
                    Notificacion.objects.bulk_create(notificaciones)
                    messages.success(request, f'Curso "{capacitacion.nombre}" creado exitosamente y se enviaron invitaciones a {len(notificaciones)} usuarios.')
                else:
                    messages.success(request, f'Curso "{capacitacion.nombre}" creado exitosamente (no se encontraron usuarios para notificar).')
            else:
                messages.success(request, f'Curso "{capacitacion.nombre}" creado exitosamente.')
                
            return redirect('capacitaciones')
        else:
            messages.error(request, 'Error al crear el curso. Por favor, revisa el formulario.')
    else:
        form = CapacitacionForm()
    
    # Lógica de Visibilidad: Superusuarios y Editores ven todo. 
    # Los demás solo ven los cursos de su compañía o a los que están invitados.
    if request.user.is_superuser or (request.user.rol and request.user.rol.editar_capacitaciones):
        capacitaciones = Capacitacion.objects.all()
    else:
        if request.user.compania:
            capacitaciones = Capacitacion.objects.filter(
                Q(creado_por=request.user) | Q(asistentes=request.user) |
                Q(companias_invitadas=request.user.compania)
            ).distinct()
        else:
            capacitaciones = Capacitacion.objects.filter(
                Q(creado_por=request.user) | Q(asistentes=request.user)
            ).distinct()

    companias = Compania.objects.all().order_by('nombre')
    usuarios_info = list(Usuario.objects.filter(is_active=True).values('id', 'compania_id'))
    usuarios_companias = {u['id']: u['compania_id'] for u in usuarios_info}

    context = {
        'form': form, 
        'capacitaciones': capacitaciones,
        'companias': companias,
        'usuarios_companias_json': json.dumps(usuarios_companias)
    }
    return render(request, 'usuarios/capacitaciones.html', context)

@login_required
def capacitacion_edit_view(request, capacitacion_id):
    capacitacion = get_object_or_404(Capacitacion, id=capacitacion_id)
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_capacitaciones):
        messages.error(request, 'No tienes permiso para editar este curso.')
        return redirect('capacitaciones')

    if request.method == 'POST':
        form = CapacitacionForm(request.POST, instance=capacitacion)
        if form.is_valid():
            capacitacion = form.save() # El save() de un ModelForm maneja las relaciones ManyToMany
            
            # Lógica de Notificaciones / Actualizaciones en Edición
            enviar_invitacion = request.POST.get('enviar_invitacion') == 'on'
            audiencia = request.POST.get('audiencia', 'general')

            if enviar_invitacion:
                if audiencia == 'general':
                    usuarios_a_notificar = Usuario.objects.filter(is_active=True, compania__in=capacitacion.companias_invitadas.all()) if capacitacion.companias_invitadas.exists() else Usuario.objects.none()
                else:
                    usuarios_a_notificar = capacitacion.asistentes.all()
                notificaciones = []
                for u in usuarios_a_notificar:
                    mensaje = f"Actualización del curso: {capacitacion.nombre}."
                    notificaciones.append(Notificacion(usuario=u, mensaje=mensaje, link=f"/capacitaciones/"))
                
                if notificaciones:
                    Notificacion.objects.bulk_create(notificaciones)
                    messages.success(request, f'Curso "{capacitacion.nombre}" actualizado y se notificó a {len(notificaciones)} usuarios.')
                else:
                    messages.success(request, f'Curso "{capacitacion.nombre}" actualizado exitosamente.')
            else:
                messages.success(request, f'Curso "{capacitacion.nombre}" actualizado exitosamente.')
        else:
            messages.error(request, 'Error al actualizar el curso. Revisa el formulario.')
    return redirect('capacitaciones')

@login_required
def capacitacion_delete_view(request, capacitacion_id):
    capacitacion = get_object_or_404(Capacitacion, id=capacitacion_id)
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_capacitaciones):
        messages.error(request, 'No tienes permiso para eliminar este curso.')
        return redirect('capacitaciones')

    if request.method == 'POST':
        capacitacion_nombre = capacitacion.nombre
        capacitacion.delete()
        messages.success(request, f'Curso "{capacitacion_nombre}" eliminado exitosamente.')
    return redirect('capacitaciones')


@login_required
def mantenimiento_view(request):
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.ver_mantenimientos):
        messages.error(request, 'No tienes permiso para acceder a Mantenimiento.')
        return redirect('dashboard')

    if request.method == 'POST':
        if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_mantenimientos):
            messages.error(request, 'No tienes permiso para registrar mantenimientos.')
            return redirect('mantenimiento')
            
        form = MantenimientoForm(request.POST)
        archivo_form = ArchivoMantenimientoForm(request.POST, request.FILES) # Para validación de archivos

        if form.is_valid() and archivo_form.is_valid():
            mantenimiento = form.save(commit=False)
            es_nuevo = not mantenimiento.pk
            
            # --- LÓGICA DE CAJA CHICA: Descuento manual ---
            caja_descuento = form.cleaned_data.get('caja_descuento')
            if es_nuevo and mantenimiento.costo > 0 and caja_descuento and caja_descuento != 'sin_costo':
                compania_caja = None if caja_descuento == 'general' else Compania.objects.filter(id=int(caja_descuento)).first()
                mantenimiento.compania_caja = compania_caja
                mantenimiento.save()
                
                CajaChica.objects.create(
                    tipo='Egreso',
                    monto=mantenimiento.costo,
                    descripcion=f"Pago por mantenimiento de {mantenimiento.vehiculo} ({mantenimiento.tipo})",
                    responsable=request.user,
                    compania=compania_caja
                )
            else:
                mantenimiento.save()
            # --------------------------------------------------

            files = request.FILES.getlist('archivos')
            for f in files:
                ArchivoMantenimiento.objects.create(mantenimiento=mantenimiento, archivo=f)

            messages.success(request, f'Mantenimiento para "{mantenimiento.vehiculo}" registrado exitosamente.')
            return redirect('mantenimiento')
        else:
            # Unir errores de ambos formularios para mostrarlos
            error_list = {**form.errors, **archivo_form.errors}
            messages.error(request, 'Error al registrar el mantenimiento. Por favor, revisa el formulario.')
    else:
        form = MantenimientoForm()
        archivo_form = ArchivoMantenimientoForm()

    # --- Lógica de Búsqueda y Filtros ---
    query = request.GET.get('q')
    tipo = request.GET.get('tipo')
    estado = request.GET.get('estado')
    usuario_id = request.GET.get('usuario')
    ordenar_por = request.GET.get('ordenar_por', '-fecha')

    # Pre-cargamos archivos y responsables para optimizar la carga de la tabla
    mantenimientos = Mantenimiento.objects.select_related('responsable').prefetch_related('archivos').all()

    if query:
        mantenimientos = mantenimientos.filter(
            Q(vehiculo__icontains=query) |
            Q(descripcion__icontains=query)
        )

    if tipo:
        mantenimientos = mantenimientos.filter(tipo=tipo)

    if estado:
        mantenimientos = mantenimientos.filter(estado=estado)

    if usuario_id:
        mantenimientos = mantenimientos.filter(responsable__id=usuario_id)

    valid_sorts = ['-fecha', 'fecha', 'vehiculo']
    if ordenar_por in valid_sorts:
        mantenimientos = mantenimientos.order_by(ordenar_por)

    usuarios_con_mantenimientos = Usuario.objects.filter(mantenimientos_responsable__isnull=False).distinct().order_by('nombre')

    context = {'form': form, 'archivo_form': archivo_form, 'mantenimientos': mantenimientos, 'usuarios_con_mantenimientos': usuarios_con_mantenimientos}
    return render(request, 'usuarios/mantenimiento.html', context)

@login_required
def mantenimiento_edit_view(request, mantenimiento_id):
    mantenimiento = get_object_or_404(Mantenimiento, id=mantenimiento_id)
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_mantenimientos):
        messages.error(request, 'No tienes permiso para editar este mantenimiento.')
        return redirect('mantenimiento')

    if request.method == 'POST':
        form = MantenimientoForm(request.POST, instance=mantenimiento)
        archivo_form = ArchivoMantenimientoForm(request.POST, request.FILES)
        if form.is_valid() and archivo_form.is_valid():
            mantenimiento_actualizado = form.save()
            
            files = request.FILES.getlist('archivos')
            for f in files:
                ArchivoMantenimiento.objects.create(mantenimiento=mantenimiento_actualizado, archivo=f)
                
            messages.success(request, f'Mantenimiento para "{mantenimiento_actualizado.vehiculo}" actualizado exitosamente.')
        else:
            messages.error(request, 'Error al actualizar el mantenimiento. Revisa el formulario.')
    return redirect('mantenimiento')

@login_required
def mantenimiento_delete_view(request, mantenimiento_id):
    mantenimiento = get_object_or_404(Mantenimiento, id=mantenimiento_id)
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_mantenimientos):
        messages.error(request, 'No tienes permiso para eliminar este mantenimiento.')
        return redirect('mantenimiento')
    if request.method == 'POST':
        mantenimiento_vehiculo = mantenimiento.vehiculo
        mantenimiento.delete()
        messages.success(request, f'Mantenimiento para "{mantenimiento_vehiculo}" eliminado exitosamente.')
    return redirect('mantenimiento')

@login_required
def caja_chica_view(request):
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.ver_caja_chica):
        messages.error(request, 'No tienes permiso para acceder a la Caja Chica.')
        return redirect('dashboard')

    if request.method == 'POST':
        if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_caja_chica):
            messages.error(request, 'No tienes permiso para registrar movimientos.')
            return redirect('caja_chica')
            
        form = CajaChicaForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            movimiento = form.save(commit=False)
            movimiento.responsable = request.user
            if not request.user.is_superuser and request.user.compania:
                movimiento.compania = request.user.compania
            movimiento.save()
            messages.success(request, f'{movimiento.get_tipo_display()} de CLP ${movimiento.monto:,.0f} registrado exitosamente.')
            return redirect('caja_chica')
        else:
            messages.error(request, 'Error al registrar el movimiento. Revisa el formulario.')
    else:
        form = CajaChicaForm(user=request.user)

    movimientos = CajaChica.objects.all()

    if not request.user.is_superuser:
        if request.user.compania:
            movimientos = movimientos.filter(compania=request.user.compania)
            nombre_caja = request.user.compania.nombre
            filtro_compania = str(request.user.compania.id)
        else:
            movimientos = CajaChica.objects.none()
            nombre_caja = "Sin Compañía"
            filtro_compania = "ninguna"
    else:
        filtro_compania = request.GET.get('compania', 'todas')
        
        if filtro_compania == 'general':
            movimientos = movimientos.filter(compania__isnull=True)
            nombre_caja = "Caja General"
        elif filtro_compania != 'todas' and filtro_compania.isdigit():
            movimientos = movimientos.filter(compania_id=filtro_compania)
            try:
                nombre_caja = Compania.objects.get(id=filtro_compania).nombre
            except:
                nombre_caja = "Caja Chica"
        else:
            nombre_caja = "Todas las Cajas (General + Compañía)"
    
    # Calcular saldo
    saldo_actual = movimientos.aggregate(
        saldo=Sum(Case(
            When(tipo='Ingreso', then=F('monto')),
            When(tipo='Egreso', then=-F('monto')),
            output_field=DecimalField()
        ))
    )['saldo'] or 0.00

    companias = Compania.objects.all() if request.user.is_superuser else (Compania.objects.filter(id=request.user.compania.id) if request.user.compania else None)

    context = {'form': form, 'movimientos': movimientos, 'saldo_actual': saldo_actual, 'companias': companias, 'filtro_compania': filtro_compania, 'nombre_caja': nombre_caja}
    return render(request, 'usuarios/caja_chica.html', context)

@login_required
def caja_chica_edit_view(request, movimiento_id):
    movimiento = get_object_or_404(CajaChica, id=movimiento_id)
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_caja_chica):
        messages.error(request, 'No tienes permiso para editar este movimiento.')
        return redirect('caja_chica')

    if request.method == 'POST':
        form = CajaChicaForm(request.POST, request.FILES, instance=movimiento, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Movimiento actualizado exitosamente.')
        else:
            messages.error(request, 'Error al actualizar el movimiento. Revisa el formulario.')
    return redirect('caja_chica')

@login_required
def caja_chica_delete_view(request, movimiento_id):
    movimiento = get_object_or_404(CajaChica, id=movimiento_id)
    if not request.user.is_superuser and not (request.user.rol and request.user.rol.editar_caja_chica):
        messages.error(request, 'No tienes permiso para eliminar este movimiento.')
        return redirect('caja_chica')
    if request.method == 'POST':
        movimiento_desc = movimiento.descripcion
        movimiento.delete()
        messages.success(request, f'Movimiento "{movimiento_desc}" eliminado exitosamente.')
    return redirect('caja_chica')

@login_required
def administracion_view(request):
    # Restringimos el acceso solo a superusuarios
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permiso para acceder a esta sección.')
        return redirect('dashboard')

    # Formularios para los modales
    user_creation_form = AdminUserCreationForm()
    # Pasamos el request para manejar errores en el formulario de edición
    user_change_form = AdminUserChangeForm(request=request)
    rol_form = RolForm()
    compania_form = CompaniaForm()

    # Búsqueda y filtrado de usuarios
    query = request.GET.get('q')
    base_users = Usuario.objects.all()

    if query:
        base_users = base_users.filter(
            Q(nombre__icontains=query) | 
            Q(apellido__icontains=query) | 
            Q(email__icontains=query)
        )

    roles = Rol.objects.all().order_by('nombre')
    companias = Compania.objects.all().order_by('nombre')
    pending_users = Usuario.objects.filter(is_active=False).order_by('-date_joined')

    # Separamos los usuarios en activos y pendientes de aprobación
    active_users = base_users.filter(is_active=True).order_by('nombre')

    context = {
        'active_users': active_users,
        'pending_users': pending_users,
        'roles': roles, 
        'companias': companias,
        'user_creation_form': user_creation_form,
        'user_change_form': user_change_form,
        'rol_form': rol_form,
        'compania_form': compania_form,
        'edit_form_user_id': request.session.pop('edit_form_user_id', None) # Para saber qué modal de edición abrir si hay error en la edición de usuario
    }
    return render(request, 'usuarios/administracion.html', context)

@login_required
def user_create_view(request):
    if not request.user.is_superuser:
        messages.error(request, 'Acción no permitida.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = AdminUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            # El RUT se guarda desde el formulario
            messages.success(request, 'Usuario creado exitosamente.')
            return redirect('administracion')
        else:
            messages.error(request, 'Error al crear el usuario. Por favor, revisa los errores en el formulario.')
            
            # Recuperar contexto para re-renderizar la página con el formulario inválido
            user_change_form = AdminUserChangeForm(request=request)
            rol_form = RolForm()
            compania_form = CompaniaForm()
            
            base_users = Usuario.objects.all()
            roles = Rol.objects.all().order_by('nombre')
            companias = Compania.objects.all().order_by('nombre')
            pending_users = Usuario.objects.filter(is_active=False).order_by('-date_joined')
            active_users = base_users.filter(is_active=True).order_by('nombre')

            context = {
                'active_users': active_users,
                'pending_users': pending_users,
                'roles': roles, 
                'companias': companias,
                'user_creation_form': form, # Mandamos el formulario con los errores
                'user_change_form': user_change_form,
                'rol_form': rol_form,
                'compania_form': compania_form,
            }
            return render(request, 'usuarios/administracion.html', context)
    return redirect('administracion')

@login_required
def user_edit_view(request, user_id):
    if not request.user.is_superuser:
        messages.error(request, 'Acción no permitida.')
        return redirect('dashboard')

    user_to_edit = get_object_or_404(Usuario, id=user_id)
    if request.method == 'POST':
        # Pasamos el request al formulario para que pueda usarlo si es necesario
        form = AdminUserChangeForm(request.POST, instance=user_to_edit, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, f'Usuario "{user_to_edit.get_full_name()}" actualizado exitosamente.')
            return redirect('administracion')
        else:
            messages.error(request, 'Error al actualizar el usuario. Por favor, revisa el formulario.')
            
            # --- Gather all context needed by administracion.html to re-render on error ---
            user_creation_form = AdminUserCreationForm()
            rol_form = RolForm()
            compania_form = CompaniaForm()
            
            base_users = Usuario.objects.all()

            roles = Rol.objects.all().order_by('nombre')
            companias = Compania.objects.all().order_by('nombre')
            pending_users = Usuario.objects.filter(is_active=False).order_by('-date_joined')
            active_users = base_users.filter(is_active=True).order_by('nombre')

            context = {
                'active_users': active_users,
                'pending_users': pending_users,
                'roles': roles, 
                'companias': companias,
                'user_creation_form': user_creation_form,
                'user_change_form': form, # Pass the invalid form instance
                'rol_form': rol_form,
                'compania_form': compania_form,
                'edit_form_user_id': user_id # Pass the user ID to re-open the modal
            }
            return render(request, 'usuarios/administracion.html', context)

    return redirect('administracion')

@login_required
def user_approve_view(request, user_id):
    if not request.user.is_superuser:
        messages.error(request, 'Acción no permitida.')
        return redirect('dashboard')

    if request.method == 'POST':
        user_to_approve = get_object_or_404(Usuario, id=user_id)
        user_to_approve.is_active = True
        user_to_approve.save()

        # Limpiar notificaciones pendientes relacionadas a este usuario para todos los admins
        Notificacion.objects.filter(
            mensaje=f"Nueva solicitud de registro: {user_to_approve.get_full_name()}",
            leida=False
        ).update(leida=True)

        messages.success(request, f'Usuario "{user_to_approve.get_full_name()}" ha sido aprobado y ahora puede iniciar sesión.')
    return redirect('administracion')

@login_required
def user_delete_view(request, user_id):
    if request.method == 'POST' and request.user.is_superuser:
        user_to_delete = get_object_or_404(Usuario, id=user_id)
        messages.success(request, f'Usuario "{user_to_delete.get_full_name()}" eliminado.')
        
        # Limpiar notificaciones pendientes relacionadas a este usuario para todos los admins
        Notificacion.objects.filter(
            mensaje=f"Nueva solicitud de registro: {user_to_delete.get_full_name()}",
            leida=False
        ).update(leida=True)
        
        user_to_delete.delete()
    return redirect('administracion')

@login_required
def rol_create_view(request):
    if not request.user.is_superuser:
        messages.error(request, 'Acción no permitida.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = RolForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Rol creado exitosamente.')
        else:
            messages.error(request, f'Error al crear el rol: {form.errors.as_text()}')
    return redirect('administracion')

@login_required
def rol_edit_view(request, rol_id):
    if not request.user.is_superuser:
        messages.error(request, 'Acción no permitida.')
        return redirect('dashboard')

    rol = get_object_or_404(Rol, id=rol_id)
    if request.method == 'POST':
        form = RolForm(request.POST, instance=rol)
        if form.is_valid():
            form.save()
            messages.success(request, f'Rol "{rol.nombre}" actualizado exitosamente.')
        else:
            messages.error(request, f'Error al actualizar el rol: {form.errors.as_text()}')
    return redirect('administracion')

@login_required
def rol_delete_view(request, rol_id):
    if request.method == 'POST' and request.user.is_superuser:
        rol = get_object_or_404(Rol, id=rol_id)
        messages.success(request, f'Rol "{rol.nombre}" eliminado.')
        rol.delete()
    return redirect('administracion')

@login_required
def compania_create_view(request):
    if not request.user.is_superuser:
        messages.error(request, 'Acción no permitida.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = CompaniaForm(request.POST, request.FILES)
        if form.is_valid():
            compania = form.save()
            
            if form.cleaned_data.get('generar_inventario'):
                items_iniciales = [
                    ('Traje estructural', 20),
                    ('Casco de bombero', 20),
                    ('Botas estructurales', 20),
                    ('Guantes de bombero', 20),
                    ('Equipo de respiración autónoma (ERA)', 15),
                    ('Manguera de 1.5 pulgadas', 10),
                    ('Manguera de 2.5 pulgadas', 5),
                    ('Pitón / Lanza de agua', 5),
                    ('Herramienta Halligan', 2),
                ]
                for item_nombre, cantidad in items_iniciales:
                    for _ in range(cantidad):
                        Inventario.objects.create(
                            nombre=item_nombre, compania=compania,
                            ubicacion='Bodega', estado='Bueno'
                        )
                messages.success(request, 'Compañía creada y su inventario base fue añadido automáticamente.')
            else:
                messages.success(request, 'Compañía creada exitosamente (sin inventario base).')
        else:
            messages.error(request, f'Error al crear la compañía: {form.errors.as_text()}')
    return redirect('administracion')

@login_required
def compania_edit_view(request, compania_id):
    if not request.user.is_superuser:
        messages.error(request, 'Acción no permitida.')
        return redirect('dashboard')

    compania = get_object_or_404(Compania, id=compania_id)
    if request.method == 'POST':
        form = CompaniaForm(request.POST, request.FILES, instance=compania)
        if form.is_valid():
            form.save()
            messages.success(request, f'Compañía "{compania.nombre}" actualizada exitosamente.')
        else:
            messages.error(request, f'Error al actualizar la compañía: {form.errors.as_text()}')
    return redirect('administracion')

@login_required
def compania_delete_view(request, compania_id):
    if request.method == 'POST' and request.user.is_superuser:
        compania = get_object_or_404(Compania, id=compania_id)
        messages.success(request, f'Compañía "{compania.nombre}" eliminada.')
        compania.delete()
    return redirect('administracion')

@login_required
def perfil_view(request):
    user = request.user
    perfil_form = PerfilForm(instance=user)
    password_form = PasswordChangeForm(user)

    if request.method == 'POST':
        if 'update_profile' in request.POST:
            perfil_form = PerfilForm(request.POST, request.FILES, instance=user)
            if perfil_form.is_valid():
                perfil_form.save()
                messages.success(request, 'Tu perfil ha sido actualizado exitosamente.')
                return redirect('perfil')
            else:
                messages.error(request, 'Por favor, corrige los errores al actualizar el perfil.')
        
        elif 'change_password' in request.POST:
            password_form = PasswordChangeForm(user, request.POST)
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)  # Importante para mantener al usuario logueado
                messages.success(request, 'Tu contraseña ha sido cambiada exitosamente.')
                return redirect('perfil')
            else:
                messages.error(request, 'Error al cambiar la contraseña. Por favor, revisa los errores.')

    capacitaciones_asistidas = user.capacitaciones_asistidas.all()
    equipos_asignados = user.inventario_asignado.all()
    
    context = {
        'perfil_form': perfil_form, 
        'password_form': password_form, 
        'user': user, 
        'capacitaciones_asistidas': capacitaciones_asistidas,
        'equipos_asignados': equipos_asignados
    }
    return render(request, 'usuarios/perfil.html', context)

@login_required
def api_notificaciones(request):
    # Filtramos para mostrar y enviar al frontend SOLO las notificaciones no leídas
    notifs = request.user.notificaciones_recibidas.filter(leida=False)[:10]
    data = [{
        'id': n.id,
        'mensaje': n.mensaje,
        'leida': n.leida,
        'link': n.link,
        'fecha': n.fecha_creacion.strftime("%d/%m %H:%M")
    } for n in notifs]
    no_leidas = request.user.notificaciones_recibidas.filter(leida=False).count()
    return JsonResponse({'notificaciones': data, 'no_leidas': no_leidas})

@login_required
def marcar_notificacion_leida(request, notif_id):
    notif = get_object_or_404(Notificacion, id=notif_id, usuario=request.user)
    notif.leida = True
    notif.save()
    if notif.link:
        return redirect(notif.link)
    return redirect('dashboard')